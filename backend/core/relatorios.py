"""
Relatórios / Exportação — gera listagens de Clientes e Processos nos
formatos Excel (.xlsx), CSV e PDF.

As funções `_gerar_excel`, `_gerar_csv` e `_gerar_pdf` são genéricas
(recebem cabeçalhos + linhas já formatadas como texto) e reutilizadas
pelos montadores específicos de cada entidade (`gerar_relatorio_clientes`
e `gerar_relatorio_processos`).
"""

import csv
import io
from datetime import datetime

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

FORMATOS_VALIDOS = ("excel", "csv", "pdf")


def _formatar_data(dt):
    if not dt:
        return "-"
    return dt.strftime("%d/%m/%Y")


# ---------------------------------------------------------------------------
# Geradores genéricos por formato
# ---------------------------------------------------------------------------

def _gerar_excel(headers, rows, titulo):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # limite do Excel para nome de aba

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    body_font = Font(name="Arial", size=10)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            valor = row[col_idx - 1]
            max_len = max(max_len, len(str(valor)) if valor is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _gerar_csv(headers, rows):
    buffer = io.StringIO()
    # utf-8-sig: garante acentuação correta ao abrir no Excel
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)

    bytes_buffer = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
    bytes_buffer.seek(0)
    return bytes_buffer


def _gerar_pdf(headers, rows, titulo, nome_escritorio):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        title=titulo,
    )

    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle(
        "Titulo",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=2,
    )
    subtitulo_style = ParagraphStyle(
        "Subtitulo",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#666666"),
        spaceAfter=14,
    )

    story = [
        Paragraph(titulo, titulo_style),
        Paragraph(
            f"{nome_escritorio} — gerado em {timezone.now().strftime('%d/%m/%Y às %H:%M')}",
            subtitulo_style,
        ),
        Spacer(1, 6),
    ]

    tabela_dados = [headers] + [[str(v) if v is not None else "-" for v in row] for row in rows]

    tabela = Table(tabela_dados, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tabela)

    doc.build(story)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Montadores específicos por entidade
# ---------------------------------------------------------------------------

_CONTENT_TYPES = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
    "pdf": "application/pdf",
}

_EXTENSOES = {
    "excel": "xlsx",
    "csv": "csv",
    "pdf": "pdf",
}


def gerar_relatorio_clientes(clientes_qs, formato, nome_escritorio):
    headers = [
        "Nome", "Tipo de Pessoa", "CPF/CNPJ", "E-mail",
        "Telefone", "Endereço", "Status", "Cadastrado em",
    ]
    rows = [
        [
            c.nome,
            "Pessoa física" if c.tipo_pessoa == "fisica" else "Pessoa jurídica",
            c.cpf_cnpj or "-",
            c.email or "-",
            c.telefone or "-",
            c.endereco or "-",
            "Ativo" if c.ativo else "Inativo",
            _formatar_data(c.criado_em),
        ]
        for c in clientes_qs
    ]
    return _montar_arquivo(headers, rows, formato, "Relatório de Clientes", nome_escritorio)


def gerar_relatorio_processos(processos_qs, formato, nome_escritorio):
    headers = [
        "Número do Processo", "Cliente", "Área Jurídica", "Status",
        "Advogado Responsável", "Comarca/Vara", "Valor da Causa",
        "Data de Distribuição", "Criado em",
    ]
    rows = [
        [
            p.numero_processo or "-",
            p.cliente.nome,
            p.get_area_juridica_display(),
            p.get_status_display(),
            p.advogado_responsavel.nome if p.advogado_responsavel_id else "-",
            p.comarca_vara or "-",
            f"R$ {p.valor_causa:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if p.valor_causa is not None else "-",
            _formatar_data(p.data_distribuicao),
            _formatar_data(p.criado_em),
        ]
        for p in processos_qs
    ]
    return _montar_arquivo(headers, rows, formato, "Relatório de Processos", nome_escritorio)


def _montar_arquivo(headers, rows, formato, titulo, nome_escritorio):
    if formato == "excel":
        return _gerar_excel(headers, rows, titulo)
    if formato == "csv":
        return _gerar_csv(headers, rows)
    if formato == "pdf":
        return _gerar_pdf(headers, rows, titulo, nome_escritorio)
    raise ValueError(f"Formato inválido: {formato}")


def content_type_para(formato):
    return _CONTENT_TYPES[formato]


def extensao_para(formato):
    return _EXTENSOES[formato]